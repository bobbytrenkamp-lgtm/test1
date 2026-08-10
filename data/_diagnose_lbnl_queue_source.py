"""TEMP diagnostic script -- finds the real, current download URL for LBNL's
"Queued Up" interconnection queue dataset and inspects its actual structure
(sheet names, header row, row count, sample rows) before any ingestion code
is written against guessed column names.

data/catalog/dataset_registry.json already flags this as "a plausible
richer source that has not been evaluated" for the interconnection_queues
dataset entry. emp.lbl.gov/eta.lbl.gov are both blocked by this sandbox's
egress proxy (confirmed via both curl and the WebFetch tool), so this must
run on a GitHub Actions runner with real network access.

Run only in CI. Deleted once real ingestion code is designed from its
output.
"""
import re
import sys
import urllib.request

CANDIDATE_PAGES = [
    "https://eta.lbl.gov/publications/us-interconnection-queue-data-0",
    "https://emp.lbl.gov/queues",
]

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; data-research-bot/1.0)"}


def fetch(url):
    req = urllib.request.Request(url, headers=HEADERS)
    with urllib.request.urlopen(req, timeout=30) as resp:
        return resp.read()


def find_xlsx_links(html_bytes, base_url):
    text = html_bytes.decode("utf-8", errors="replace")
    # Absolute and relative .xlsx hrefs
    hrefs = re.findall(r'href="([^"]+\.xlsx[^"]*)"', text, re.IGNORECASE)
    out = []
    for h in hrefs:
        if h.startswith("http"):
            out.append(h)
        elif h.startswith("/"):
            from urllib.parse import urlparse
            p = urlparse(base_url)
            out.append(f"{p.scheme}://{p.netloc}{h}")
    return sorted(set(out))


def main():
    all_links = []
    for page in CANDIDATE_PAGES:
        print(f"\n{'=' * 70}\nFetching: {page}\n{'=' * 70}")
        try:
            html = fetch(page)
        except Exception as e:
            print(f"  FAILED to fetch page: {e}")
            continue
        print(f"  page size: {len(html)} bytes")
        links = find_xlsx_links(html, page)
        print(f"  .xlsx links found: {links}")
        all_links.extend(links)

    all_links = sorted(set(all_links))
    if not all_links:
        print("\nNo .xlsx links found on either page. Aborting inspection.")
        sys.exit(1)

    xlsx_url = all_links[0]
    print(f"\n{'=' * 70}\nDownloading: {xlsx_url}\n{'=' * 70}")
    try:
        data = fetch(xlsx_url)
    except Exception as e:
        print(f"FAILED to download xlsx: {e}")
        sys.exit(1)
    print(f"downloaded {len(data)} bytes")

    with open("/tmp/queued_up.xlsx", "wb") as f:
        f.write(data)

    import openpyxl
    wb = openpyxl.load_workbook("/tmp/queued_up.xlsx", read_only=True, data_only=True)
    print(f"\nsheet names ({len(wb.sheetnames)}): {wb.sheetnames}")

    # Look for the project-level raw-data sheet specifically (usually the
    # first sheet, or one with "data" in the name -- not a summary/codebook
    # tab).
    candidates = [s for s in wb.sheetnames if re.search(r"data|queue|project", s, re.IGNORECASE)]
    sheet_name = candidates[0] if candidates else wb.sheetnames[0]
    print(f"inspecting sheet: {sheet_name!r}")
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    print(f"\nheader row ({len(header) if header else 0} columns):")
    print(header)

    print("\nfirst 3 data rows:")
    for i, row in enumerate(rows_iter):
        if i >= 3:
            break
        print(row)

    row_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in _))
    print(f"\napprox non-empty data rows: {row_count}")


if __name__ == "__main__":
    main()
