#!/usr/bin/env python3
"""data/national_data_ingestion/interconnection_queue.py — LBNL "Queued Up"
generator interconnection queue -> InfrastructureAsset records
(asset_type="interconnection_queue_entry").

WHAT THIS IS
LBNL (Lawrence Berkeley National Laboratory) compiles, cleans, and publishes
the interconnection request queues from 7 ISO/RTOs and 50+ non-ISO
balancing areas -- ~98% of US generating capacity -- as one project-level
workbook, updated periodically (emp.lbl.gov/queues). This is the "who's
trying to connect to the grid, where, how big, and how far along" signal
data/catalog/dataset_registry.json already flagged as unevaluated for the
interconnection_queues dataset entry.

REAL SOURCE STRUCTURE (verified via a live GitHub Actions dispatch,
2026-08-11 -- never guessed): the workbook's '03. Complete Queue Data'
sheet has 30 columns and ~38,000 rows. RAW_HEADER below is the exact,
real header row.

DOWNLOAD REQUIRES A REAL BROWSER
emp.lbl.gov is behind Cloudflare's managed challenge; see
interconnection_queue_download.mjs's header for the full story. That
script writes the raw .xlsx to disk; this module never fetches anything
itself -- it only parses an already-downloaded file, so its logic is
fully testable with a fixture workbook and no network access.

GEOMETRY: COUNTY-LEVEL, NOT PER-PROJECT
The source publishes county + state text, never per-project coordinates.
Locations here are the real county's bbox centroid (see
lib/county_geometry.py, decoded from the same vendored Census TIGER
boundaries this project's own map already renders) -- an honest, derived
approximation, always tagged evidence_tier="MODELED", never "OBSERVED".
A row whose county cannot be resolved to a real FIPS centroid is excluded
and counted in the output's `excluded` block, never silently dropped.

QUEUE STATUS IS ITS OWN VOCABULARY
q_status ("active"/"withdrawn"/"operational"/...) is preserved verbatim in
`queue_status`, NOT mapped onto InfrastructureAsset's base `status` field
(existing/planned/retired/...) -- "withdrawn" (never built) and "retired"
(built, then decommissioned) are different facts; forcing one into the
other would misrepresent real project history. See
data/infrastructure_asset_schema.py's interconnection_queue_entry
TYPE_SCHEMAS comment for the same reasoning.

Usage:
  node data/national_data_ingestion/interconnection_queue_download.mjs /tmp/lbnl_queue.xlsx
  python3 data/national_data_ingestion/interconnection_queue.py /tmp/lbnl_queue.xlsx
  python3 data/national_data_ingestion/interconnection_queue.py /tmp/lbnl_queue.xlsx --check
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT))

from data.national_data_ingestion.lib.county_geometry import county_centroid  # noqa: E402

OUTPUT_PATH = ROOT / "data" / "interconnection_queue.json"
SHEET_NAME = "03. Complete Queue Data"
SOURCE_URL = "https://emp.lbl.gov/queues"
SOURCE_PUBLISHER = "Lawrence Berkeley National Laboratory (LBNL)"

# The real column order in the source sheet, confirmed via live dispatch
# 2026-08-11 -- never guessed. Column count and order both matter: this
# module reads rows positionally (openpyxl's read-only iterator has no
# named-column API), so a source schema change would need this constant
# updated deliberately, not silently misaligned.
RAW_HEADER = (
    "q_id", "q_status", "q_date", "prop_date", "on_date", "wd_date", "ia_date",
    "IA_phase_raw", "IA_phase_clean", "county", "state", "fips_code", "poi_name",
    "region", "project_name", "utility", "entity", "developer", "cluster",
    "service", "project_type", "type_1", "type_2", "type_3", "type_clean",
    "mw_1", "mw_2", "mw_3", "q_year", "prop_year",
)


def _iso_date(v: Any) -> Optional[str]:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, str) and v.strip():
        return v.strip()
    return None


def _sum_mw(*values: Any) -> Optional[float]:
    """Sums whichever of mw_1/mw_2/mw_3 are real numbers -- a hybrid project
    (e.g. solar + storage) reports its components across multiple columns;
    a single-technology project only ever populates mw_1. Returns None
    (never 0) when nothing is a real number, since "no capacity reported"
    and "reported as zero" are different facts."""
    total = None
    for v in values:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            total = (total or 0) + v
    return total


def parse_row(row: tuple) -> Optional[dict]:
    """Pure: one raw row tuple (RAW_HEADER order) -> one InfrastructureAsset
    dict, or None if the row has no usable identity (no q_status at all --
    a genuinely blank/malformed row, not a real record). Never raises on a
    single bad row; a batch caller should still isolate failures per-row.
    """
    if len(row) < len(RAW_HEADER):
        row = row + (None,) * (len(RAW_HEADER) - len(row))
    r = dict(zip(RAW_HEADER, row))

    q_status = r.get("q_status")
    if not isinstance(q_status, str) or not q_status.strip():
        return None

    fips_raw = r.get("fips_code")
    fips = None
    if isinstance(fips_raw, (int, float)) and not isinstance(fips_raw, bool):
        fips = str(int(fips_raw)).zfill(5)
    elif isinstance(fips_raw, str) and fips_raw.strip().isdigit():
        fips = fips_raw.strip().zfill(5)

    centroid = county_centroid(fips) if fips else None

    name = (
        (r.get("project_name") or "").strip()
        or (r.get("poi_name") or "").strip()
        or f"Interconnection queue entry {r.get('q_id') or 'unassigned'}"
    )

    asset: dict = {
        # Row position is folded into the id (via the caller in
        # parse_workbook_rows) because q_id alone is not always unique or
        # even present ("not assigned" appears verbatim in real source
        # rows) -- see this module's own header for why a stable natural
        # key isn't available here.
        "asset_type": "interconnection_queue_entry",
        "name": name,
        "source": {
            "publisher": SOURCE_PUBLISHER,
            "url": SOURCE_URL,
            "retrieved_at": None,  # filled in by parse_workbook()
        },
        "evidence_tier": "MODELED" if centroid else "UNKNOWN",
        "last_verified": None,  # filled in by parse_workbook()
        "queue_status": q_status.strip(),
    }

    if centroid:
        asset["geometry"] = {"type": "Point", "coordinates": centroid}
    else:
        asset["geometry"] = None

    capacity_mw = _sum_mw(r.get("mw_1"), r.get("mw_2"), r.get("mw_3"))
    if capacity_mw is not None:
        asset["capacity_mw"] = capacity_mw
    if r.get("type_clean"):
        asset["technology"] = r["type_clean"]
    if r.get("utility"):
        asset["utility"] = r["utility"]
    if r.get("poi_name"):
        asset["point_of_interconnection"] = r["poi_name"]
    if r.get("q_id"):
        asset["queue_id"] = r["q_id"]
    if r.get("developer"):
        asset["developer"] = r["developer"]
    if _iso_date(r.get("q_date")):
        asset["queue_date"] = _iso_date(r.get("q_date"))
    if _iso_date(r.get("prop_date")):
        asset["proposed_online_date"] = _iso_date(r.get("prop_date"))
    if _iso_date(r.get("wd_date")):
        asset["withdrawn_date"] = _iso_date(r.get("wd_date"))
    if _iso_date(r.get("ia_date")):
        asset["interconnection_agreement_date"] = _iso_date(r.get("ia_date"))
    if r.get("service"):
        asset["service_type"] = r["service"]
    if r.get("region"):
        asset["balancing_authority_region"] = r["region"]
    if r.get("county") and r.get("state"):
        asset["county_state"] = f"{r['county']}, {r['state']}"
    if fips:
        asset["county_fips"] = fips

    return asset


def parse_workbook_rows(rows: list, retrieved_at: str) -> dict:
    """Pure (given already-extracted rows, no file I/O): builds the full
    output document. `rows` excludes the header. Every row lands in exactly
    one bucket -- assets (usable) or excluded (no q_status at all, or no
    resolvable location) -- so total accounting is always exact, never a
    silent drop."""
    assets = []
    excluded_no_status = 0
    excluded_no_location = 0

    for i, row in enumerate(rows):
        asset = parse_row(row)
        if asset is None:
            excluded_no_status += 1
            continue
        asset["id"] = f"lbnl-iq:{i}:{asset.get('queue_id') or 'na'}"
        asset["source"]["retrieved_at"] = retrieved_at
        asset["last_verified"] = retrieved_at
        if asset["geometry"] is None:
            excluded_no_location += 1
            continue
        assets.append(asset)

    return {
        "meta": {
            "generated_at": retrieved_at,
            "source_publisher": SOURCE_PUBLISHER,
            "source_url": SOURCE_URL,
            "total_rows_in_source": len(rows),
            "included": len(assets),
            "excluded_no_status": excluded_no_status,
            "excluded_no_resolvable_county_location": excluded_no_location,
            "caveat": (
                "Locations are county-level bbox centroids (evidence_tier=MODELED), "
                "not per-project coordinates -- the source publishes county/state "
                "text only, never exact site coordinates. Records whose county could "
                "not be resolved to a real FIPS centroid are excluded and counted "
                "above, not silently dropped. queue_status uses the source's own "
                "vocabulary (e.g. active/withdrawn/operational), distinct from this "
                "project's generic existing/planned/retired status values."
            ),
        },
        "assets": assets,
    }


def parse_workbook(xlsx_path: Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"expected sheet {SHEET_NAME!r} not found; workbook has: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    rows_iter = ws.iter_rows(values_only=True)
    next(rows_iter, None)  # 'RETURN TO CONTENTS' banner row
    header = next(rows_iter, None)
    if header is None or tuple(header[: len(RAW_HEADER)]) != RAW_HEADER:
        raise ValueError(f"unexpected header row: {header!r} -- source schema may have changed")

    rows = [row for row in rows_iter if any(c is not None for c in row)]
    retrieved_at = datetime.utcnow().strftime("%Y-%m-%d")
    return parse_workbook_rows(rows, retrieved_at)


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("xlsx_path", type=Path)
    p.add_argument("--check", action="store_true", help="verify the committed output matches a fresh parse, don't write")
    p.add_argument("--output", type=Path, default=OUTPUT_PATH)
    args = p.parse_args()

    if not args.xlsx_path.exists():
        print(f"FATAL: {args.xlsx_path} does not exist", file=sys.stderr)
        return 2

    doc = parse_workbook(args.xlsx_path)

    if args.check:
        if not args.output.exists():
            print(f"FAILED: {args.output} does not exist")
            return 1
        committed = json.loads(args.output.read_text())
        # generated_at legitimately differs run-to-run; compare everything else.
        fresh_assets = doc["assets"]
        committed_assets = committed.get("assets", [])
        if fresh_assets != committed_assets:
            print(f"FAILED: {args.output} does not match a fresh parse of {args.xlsx_path}")
            return 1
        print(f"OK -- {args.output} matches a fresh parse ({len(fresh_assets)} assets).")
        return 0

    args.output.write_text(json.dumps(doc, indent=2) + "\n")
    m = doc["meta"]
    print(
        f"Wrote {args.output}: {m['included']} assets "
        f"({m['excluded_no_status']} excluded no-status, "
        f"{m['excluded_no_resolvable_county_location']} excluded no-location) "
        f"from {m['total_rows_in_source']} source rows."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
