"""FERC Interconnection Queue adapter.

The Federal Energy Regulatory Commission publishes the active interconnection
request queues for all US transmission providers. Large power requests (>50 MW)
from commercial/industrial applicants in the queue frequently correspond to
planned data center developments.

Public source (no auth required):
  FERC Queue download portal: https://www.ferc.gov/industries-data/electric/
                               electric-power-markets/interconnection-queues

Each ISO/RTO publishes its own queue. FERC aggregates them; individual ISOs
also publish directly:
  PJM:  https://www.pjm.com/planning/interconnection-projects
  MISO: https://www.misoenergy.org/planning/interconnection/
  CAISO: https://www.caiso.com/planning/Pages/GeneratorInterconnection/
  SPP:  https://www.spp.org/engineering/generator-interconnection/

Tier: 4 (discovery) — power requests confirm development intent but operator,
exact location, and facility type must be verified via other sources.
All records from this source are marked as candidates.

Note: The FERC master file is a public Excel/CSV download. Parse it with
openpyxl or pandas. Filter for:
  - Power request >= 50 MW (data centers rarely request less)
  - Project type: "Synchronous Condenser", "Storage", or blank (grid load)
  - Status: "Active", "In Progress", or "Queue"
"""
from __future__ import annotations

import re
import time
from typing import Iterator

from ..models import FacilityRecord, FacilitySource
from ..normalize import normalize_record_fields, normalize_state
from . import BaseAdapter

# FERC master queue download — public Excel file, updated periodically
FERC_QUEUE_URL = (
    "https://www.ferc.gov/media/3423/download"
)

# Per-ISO active queue download URLs (updated periodically by each ISO)
ISO_QUEUE_URLS: dict[str, str] = {
    "PJM": "https://www.pjm.com/-/media/planning/interconnection-projects/active-projects/ai-queues.ashx",
    "MISO": "https://www.misoenergy.org/planning/interconnection/disis/",
    "CAISO": "https://www.caiso.com/Documents/GeneratorInterconnectionandInterconnectionRelatedAgreementsReport.xlsx",
    "SPP": "https://www.spp.org/documents/49169/SPP_GI_Queue.xlsx",
    "NYISO": "https://www.nyiso.com/documents/20142/1407078/New-York-ISO-Active-Interconnection-Queue.xlsx",
    "ISO-NE": "https://www.iso-ne.com/system-planning/interconnection-service/interconnection-request-queue/current-request-queue",
    "ERCOT": "https://mis.ercot.com/misapp/GetReports.do?reportTypeId=15933",
}

# MW threshold for candidate data center projects
_MIN_MW = 50

# Keywords that suggest a data center in the project name / applicant name
_DC_KEYWORDS = [
    "data center", "datacenter", "cloud", "colocation", "colo",
    "hyperscale", "computing", "artificial intelligence", "ai campus",
    "server", "digital", "technology campus",
]


def _row_to_record(row: dict, source_url: str, source_id: str) -> FacilityRecord | None:
    """Convert one FERC queue row to a FacilityRecord.

    Row keys vary by ISO; this tries common column names.
    """
    # Capacity — look for common column names
    mw_raw = (
        row.get("MW (Summer)") or row.get("MW") or row.get("Capacity (MW)")
        or row.get("Requested MW") or row.get("Net MW") or ""
    )
    try:
        mw = float(str(mw_raw).replace(",", "").strip())
    except (TypeError, ValueError):
        mw = 0.0

    if mw < _MIN_MW:
        return None

    # Project / applicant name
    project_name = (
        row.get("Project Name") or row.get("Queue Project Name")
        or row.get("Interconnection Customer") or row.get("Developer") or ""
    ).strip()

    applicant = (
        row.get("Applicant") or row.get("Developer") or
        row.get("Interconnection Customer") or project_name
    ).strip()

    # Heuristic: does the name suggest a data center?
    combined_text = (project_name + " " + applicant).lower()
    is_dc_hint = any(kw in combined_text for kw in _DC_KEYWORDS)
    if not is_dc_hint and mw < 100:
        # Only include smaller projects if name suggests DC
        return None

    # Location
    state_raw = (
        row.get("State") or row.get("state") or row.get("Location State") or ""
    ).strip()
    county_raw = (
        row.get("County") or row.get("county") or ""
    ).strip()
    full_state, state_abbr = normalize_state(state_raw)
    if not state_abbr:
        return None  # Can't use record without state

    # Status
    status_raw = (
        row.get("Status") or row.get("Queue Status") or row.get("Project Status") or "planned"
    ).strip().lower()
    if "withdrawn" in status_raw or "terminated" in status_raw:
        return None
    operational_status = "planned"
    if "operational" in status_raw or "in service" in status_raw:
        operational_status = "operational"
    elif "construction" in status_raw or "under construction" in status_raw:
        operational_status = "under_construction"

    # Date
    queue_date = (
        row.get("Queue Date") or row.get("Submitted Date") or ""
    ).strip()

    r = FacilityRecord()
    r.name = project_name or f"{applicant} Data Center"
    r.operator = applicant
    r.state = full_state
    r.state_abbr = state_abbr
    r.county = county_raw
    r.operational_status = operational_status
    r.capacity_mw_planned = mw
    r.primary_source = source_id
    r.confidence_tier = 4
    r.confidence_score = 0.40
    r.is_candidate = True
    r.last_verified_date = queue_date[:10] if len(queue_date) >= 10 else queue_date

    if source_url:
        r.source_urls.append(source_url)

    r.notes = (
        f"FERC interconnection queue request: {mw:.0f} MW in {state_abbr}"
        + (f", {county_raw} County" if county_raw else "")
        + f". Queue date: {queue_date}. "
        + "Needs verification that this is a data center project."
    )

    normalize_record_fields(r)
    return r


def _fetch_ferc_master(session) -> list[dict]:
    """Download and parse the FERC master interconnection queue Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required: pip install openpyxl")

    import io
    try:
        time.sleep(1.0)
        resp = session.get(FERC_QUEUE_URL, timeout=60)
        resp.raise_for_status()
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
        return rows
    except Exception:
        return []


class FERCQueueAdapter(BaseAdapter):
    """Discovers large power interconnection requests that may be data center projects.

    All records are candidates (tier 4, confidence 0.40) and require human
    verification before promotion to master. Large >100 MW requests with
    DC-related names are prioritized.
    """

    def __init__(self, source: FacilitySource):
        super().__init__(source)

    def fetch(self, since: str | None = None) -> Iterator[FacilityRecord]:
        try:
            import requests
        except ImportError:
            raise RuntimeError("requests required: pip install requests")

        session = requests.Session()
        session.headers.update({
            "User-Agent": "US-AI-Infrastructure-Map/1.0 datacenter-research@example.com",
        })

        rows = _fetch_ferc_master(session)
        seen: set[str] = set()

        for row in rows:
            record = _row_to_record(row, FERC_QUEUE_URL, self.source_id)
            if record is None:
                continue
            # Deduplicate within this batch by name+state
            key = f"{record.name}|{record.state_abbr}"
            if key in seen:
                continue
            seen.add(key)
            yield self._stamp(record)
