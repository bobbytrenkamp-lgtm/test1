"""TEMP: inspects /tmp/queued_up.xlsx (downloaded by
_diagnose_lbnl_queue_browser.mjs) -- sheet names, header row, sample rows,
approx row count. See that script's header for full context. Deleted once
real ingestion code is designed from this output.
"""
import re

import openpyxl

wb = openpyxl.load_workbook("/tmp/queued_up.xlsx", read_only=True, data_only=True)
print(f"\nsheet names ({len(wb.sheetnames)}): {wb.sheetnames}")

# Round 4 revealed the real sheet list: '02. Data Sample by Region' is a
# small summary tab that also happens to match a loose "data" regex, while
# '03. Complete Queue Data' is the actual project-level record sheet.
# Prefer an exact/near match on "complete queue data" first.
exact = [s for s in wb.sheetnames if re.search(r"complete.*queue.*data", s, re.IGNORECASE)]
candidates = exact or [s for s in wb.sheetnames if re.search(r"data|queue|project", s, re.IGNORECASE)]
sheet_name = candidates[0] if candidates else wb.sheetnames[0]
print(f"inspecting sheet: {sheet_name!r}")
ws = wb[sheet_name]

rows_iter = ws.iter_rows(values_only=True)
header = next(rows_iter, None)
print(f"\nheader row ({len(header) if header else 0} columns):")
print(header)

print("\nfirst 5 data rows:")
for i, row in enumerate(rows_iter):
    if i >= 5:
        break
    print(row)

row_count = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r))
print(f"\napprox non-empty data rows: {row_count}")
